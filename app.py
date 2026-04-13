import io
import os
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Page config ───────────────────────────────
st.set_page_config(
    page_title="Resource Report Generator",
    page_icon="📊",
    layout="centered",
)

# ── Constants ─────────────────────────────────
EXCLUDED_DEPARTMENTS = {
    'Administration', 'Collaboration Services', 'Digital Marketing',
    'Finance', 'HR', 'IT System Engineering', 'Maintenance',
    'Management', 'Products and Services', 'Salesforce',
}

KEYWORD_RULES = [
    ('highsystem', 'Nexus Schweiz AG'),
    ('star',       'STAR Enterprises AG'),
    ('sky',        'STAR Enterprises AG'),
]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ── Core logic ────────────────────────────────

def load_employee_map():
    path = os.path.join(SCRIPT_DIR, 'Employee.xlsx')
    if not os.path.exists(path):
        return {}, {}
    df = pd.read_excel(path, header=None)
    role_map, location_map = {}, {}
    for i, row in df.iterrows():
        vals = [str(v).strip() if pd.notna(v) else '' for v in row]
        if 'Employee Name' in vals and 'Role' in vals:
            ec = vals.index('Employee Name')
            rc = vals.index('Role')
            lc = vals.index('Location') if 'Location' in vals else None
            for _, dr in df.iloc[i + 1:].iterrows():
                emp  = str(dr.iloc[ec]).strip() if pd.notna(dr.iloc[ec]) else ''
                role = str(dr.iloc[rc]).strip() if pd.notna(dr.iloc[rc]) else ''
                loc  = str(dr.iloc[lc]).strip() if (lc is not None and pd.notna(dr.iloc[lc])) else ''
                if emp and emp.lower() != 'nan':
                    role_map[emp.lower()]     = role
                    location_map[emp.lower()] = loc
            break
    return role_map, location_map


def load_client_map():
    path = os.path.join(SCRIPT_DIR, 'Project.xlsx')
    if not os.path.exists(path):
        return {}
    df = pd.read_excel(path, header=None)
    client_map = {}
    for i, row in df.iterrows():
        vals = [str(v).strip() if pd.notna(v) else '' for v in row]
        if 'Project Name' in vals and 'Client' in vals:
            pc = vals.index('Project Name')
            cc = vals.index('Client')
            for _, dr in df.iloc[i + 1:].iterrows():
                proj   = str(dr.iloc[pc]).strip() if pd.notna(dr.iloc[pc]) else ''
                client = str(dr.iloc[cc]).strip() if pd.notna(dr.iloc[cc]) else ''
                if proj and proj.lower() != 'nan':
                    client_map[proj.lower()] = client
            break
    return client_map


def lookup_client(proj_name, client_map):
    key = proj_name.lower()
    if key in client_map:
        return client_map[key]
    for map_key, client in client_map.items():
        if map_key in key or key in map_key:
            return client
    for keyword, client in KEYWORD_RULES:
        if keyword in key:
            return client
    return ''


def identify_departments(df):
    not_null_idx = df[df[0].notna()].index.tolist()
    dept_set = set()
    for i, idx in enumerate(not_null_idx):
        val = df.iloc[idx][0]
        if val in ['MIS Report', 'Employees']:
            continue
        next_idx = not_null_idx[i + 1] if i + 1 < len(not_null_idx) else None
        if next_idx and next_idx == idx + 1:
            dept_set.add(val)
    return dept_set


def parse_data(df, dept_set):
    current_dept, current_emp = None, None
    emp_to_record = {}
    for idx in range(len(df)):
        row = df.iloc[idx]
        col0, col1, col2, col3 = row[0], row[1], row[2], row[3]
        if pd.notna(col0):
            val = col0
            if val in ['MIS Report', 'Employees']:
                continue
            if val in dept_set:
                current_dept = val
                current_emp  = None
            else:
                current_emp = val
                if current_dept and current_dept not in EXCLUDED_DEPARTMENTS:
                    key = (current_dept, current_emp)
                    if key not in emp_to_record:
                        emp_to_record[key] = {
                            'dept': current_dept, 'employee': current_emp,
                            'pr_projects': set(), 'ipr_projects': set(),
                        }
        elif pd.notna(col1) and pd.notna(col2) and pd.notna(col3):
            proj_num  = str(col2).strip()
            proj_name = str(col3).strip()
            if current_emp and current_dept and current_dept not in EXCLUDED_DEPARTMENTS:
                key = (current_dept, current_emp)
                if key in emp_to_record:
                    if proj_num.startswith('PR'):
                        emp_to_record[key]['pr_projects'].add((proj_num, proj_name))
                    elif proj_num.startswith('IPR'):
                        emp_to_record[key]['ipr_projects'].add((proj_num, proj_name))
    return emp_to_record


def build_output_rows(emp_to_record, role_map, location_map, client_map):
    output_rows = []
    for (dept, emp), rec in emp_to_record.items():
        pr  = sorted(rec['pr_projects'])
        ipr = sorted(rec['ipr_projects'])
        if pr:
            projects = [{'name': p[1], 'billable': 'Yes'} for p in pr]
        elif ipr:
            projects = [{'name': ipr[0][1], 'billable': 'No'}]
        else:
            continue
        allocation = 'Part Time' if len(projects) > 1 else 'Full Time'
        role       = role_map.get(emp.lower(), '')
        location   = location_map.get(emp.lower(), '')
        for proj in projects:
            client = lookup_client(proj['name'], client_map)
            output_rows.append({
                'Department':     dept,
                'Employee':       emp,
                'Role':           role,
                'Location':       location,
                'Project Name':   proj['name'],
                'Client':         client,
                'Allocation':     allocation,
                'Billable (Y/N)': proj['billable'],
            })
    output_rows.sort(key=lambda x: (x['Department'], x['Employee'], x['Project Name']))
    return output_rows


def build_excel_bytes(output_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Resource Report"

    headers     = ['Department', 'Employee', 'Role', 'Location', 'Project Name', 'Client', 'Allocation', 'Billable (Y/N)']
    hfill       = PatternFill('solid', start_color='1F3864')
    hfont       = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[1].height = 30

    dept_list   = sorted(set(r['Department'] for r in output_rows))
    palette     = ['EBF3FB', 'FEF9E7', 'EAFAF1', 'FDF2F8', 'F4ECF7', 'FDEDEC', 'E8F8F5', 'FEF5E7', 'EAF2FF']
    dept_colors = {d: palette[i % len(palette)] for i, d in enumerate(dept_list)}
    alloc_colors    = {'Full Time': 'D5F5E3', 'Part Time': 'FEF9E7'}
    billable_colors = {'Yes': 'D5F5E3', 'No': 'FDEDEC'}

    row_num = 2
    for r in output_rows:
        dept     = r['Department']
        row_fill = PatternFill('solid', start_color=dept_colors[dept])
        for col, val in enumerate([dept, r['Employee'], r['Role'], r['Location'], r['Project Name'], r['Client']], 1):
            cell           = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = row_fill
            cell.font      = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(col == 5))
            cell.border    = thin_border
        av   = r['Allocation']
        cell = ws.cell(row=row_num, column=7, value=av)
        cell.fill      = PatternFill('solid', start_color=alloc_colors.get(av, 'FFFFFF'))
        cell.font      = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        bv   = r['Billable (Y/N)']
        cell = ws.cell(row=row_num, column=8, value=bv)
        cell.fill      = PatternFill('solid', start_color=billable_colors.get(bv, 'FFFFFF'))
        cell.font      = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        row_num += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.freeze_panes    = 'A2'
    ws.auto_filter.ref = f"A1:H{row_num - 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), row_num - 2, len(dept_list)


# ── UI ────────────────────────────────────────

st.title("📊 Resource Report Generator")
st.markdown("Upload the weekly MIS Report and download the consolidated output instantly.")
st.divider()

# Load mapping files (cached so they don't reload on every interaction)
@st.cache_data
def get_maps():
    role_map, location_map = load_employee_map()
    client_map             = load_client_map()
    return role_map, location_map, client_map

role_map, location_map, client_map = get_maps()

# Show mapping status
col1, col2 = st.columns(2)
with col1:
    if role_map:
        st.success(f"✅ Employee.xlsx — {len(role_map)} employees")
    else:
        st.error("❌ Employee.xlsx not found on server")
with col2:
    if client_map:
        st.success(f"✅ Project.xlsx — {len(client_map)} projects")
    else:
        st.error("❌ Project.xlsx not found on server")

st.divider()

# File uploader
uploaded_file = st.file_uploader(
    "Upload your MIS Report (.xlsx)",
    type=["xlsx"],
    help="This is the weekly export from your MIS system.",
)

if uploaded_file:
    st.info(f"📄 Uploaded: **{uploaded_file.name}**")

    if st.button("⚙️ Generate Consolidated Report", type="primary", use_container_width=True):
        with st.spinner("Processing..."):
            try:
                df            = pd.read_excel(uploaded_file, header=None, sheet_name=0)
                dept_set      = identify_departments(df)
                emp_to_record = parse_data(df, dept_set)
                output_rows   = build_output_rows(emp_to_record, role_map, location_map, client_map)

                if not output_rows:
                    st.error("No data found. Please check the uploaded file.")
                else:
                    excel_bytes, total_rows, total_depts = build_excel_bytes(output_rows)

                    st.success(f"✅ Done! **{total_rows} rows** across **{total_depts} departments**.")

                    # Output filename based on input filename
                    base        = os.path.splitext(uploaded_file.name)[0]
                    output_name = f"Consolidated_{base}.xlsx"

                    st.download_button(
                        label="⬇️ Download Report",
                        data=excel_bytes,
                        file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )

            except Exception as e:
                st.error(f"Something went wrong: {e}")

st.divider()
st.caption("Employee.xlsx and Project.xlsx are managed by the admin on the server. Contact your admin to update mappings.")
